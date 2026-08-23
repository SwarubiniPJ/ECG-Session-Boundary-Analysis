from __future__ import annotations

import json
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .utils import ensure_dir


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    caution_fill = PatternFill("solid", fgColor="FFF2CC")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_index, column_cells in enumerate(sheet.columns, start=1):
            values = [str(cell.value) for cell in column_cells[:200] if cell.value is not None]
            width = min(max([len(value) for value in values] + [8]) + 2, 42)
            sheet.column_dimensions[get_column_letter(column_index)].width = width
        headers = {cell.value: cell.column for cell in sheet[1]}
        if "Status" in headers:
            column = headers["Status"]
            for row in range(2, sheet.max_row + 1):
                if str(sheet.cell(row, column).value).lower() not in {"ok", "complete", "true"}:
                    sheet.cell(row, column).fill = caution_fill
    workbook.save(path)


def write_excel_tables(path: Path, tables: dict[str, pd.DataFrame]) -> None:
    ensure_dir(path.parent)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, table in tables.items():
            table.to_excel(writer, sheet_name=name[:31], index=False)
    style_workbook(path)


def write_paper_tables(
    output_dir: Path,
    tables: dict[str, pd.DataFrame],
) -> Path:
    ensure_dir(output_dir)
    clean_tables: dict[str, pd.DataFrame] = {}
    for name, table in tables.items():
        clean = table.copy() if table is not None else pd.DataFrame()
        clean_tables[name] = clean
        clean.to_csv(output_dir / f"{name}.csv", index=False)
    workbook = output_dir / "Nature_Timing_Paper_Tables_V4.xlsx"
    write_excel_tables(workbook, clean_tables)
    return workbook


def _save_figure(fig: plt.Figure, path: Path, dpi: int) -> None:
    ensure_dir(path.parent)
    fig.savefig(path.with_suffix(".png"), dpi=dpi, bbox_inches="tight", facecolor="white")
    fig.savefig(path.with_suffix(".pdf"), bbox_inches="tight", facecolor="white")
    plt.close(fig)


def make_figures(
    output_dir: Path,
    evaluation: pd.DataFrame,
    detection_summary: pd.DataFrame,
    comparisons: pd.DataFrame,
    simulation_power: pd.DataFrame,
    boundary_order: pd.DataFrame,
    quality_eligibility: pd.DataFrame,
    feature_summary: pd.DataFrame,
    empirical_pvalues: pd.DataFrame,
    dpi: int,
    population_timing_summary: pd.DataFrame | None = None,
    population_timing_profiles: pd.DataFrame | None = None,
    population_timing_simulation: pd.DataFrame | None = None,
) -> None:
    ensure_dir(output_dir)
    plt.rcParams.update(
        {
            "font.family": "sans-serif",
            "font.sans-serif": ["Arial", "Helvetica", "DejaVu Sans"],
            "font.size": 8,
            "axes.spines.top": False,
            "axes.spines.right": False,
            "pdf.fonttype": 42,
        }
    )

    primary = detection_summary[
        detection_summary["RRThreshold"].eq(20.0)
        & detection_summary["Representation"].eq("reduced")
        & detection_summary["SearchWindow"].eq("post_only")
        & detection_summary["Method"].eq("CovIC_crossfit")
    ].copy()
    if not primary.empty:
        fig, ax = plt.subplots(figsize=(6.5, 3.8))
        x = np.arange(len(sorted(primary["WindowLength_sec"].unique())))
        windows = sorted(primary["WindowLength_sec"].unique())
        groups = [
            ("Real", "", "Real"),
            ("Pseudo", "A", "Pseudo within A"),
            ("Pseudo", "NA", "Pseudo within NA"),
        ]
        offsets = [-0.18, 0.0, 0.18]
        for offset, (kind, condition, label) in zip(offsets, groups):
            subset = primary[
                primary["BoundaryKind"].eq(kind)
                & primary["PseudoCondition"].fillna("").eq(condition)
            ].groupby("WindowLength_sec", as_index=False).agg(
                DetectionRate=("DetectionRate", "mean"),
                Low=("DetectionCI95_Lower", "mean"),
                High=("DetectionCI95_Upper", "mean"),
            ).set_index("WindowLength_sec").reindex(windows)
            y = 100 * subset["DetectionRate"].to_numpy(dtype=float)
            low = 100 * subset["Low"].to_numpy(dtype=float)
            high = 100 * subset["High"].to_numpy(dtype=float)
            ax.errorbar(
                x + offset,
                y,
                yerr=np.vstack([y - low, high - y]),
                marker="o",
                capsize=3,
                label=label,
            )
        ax.set_xticks(x, windows)
        ax.set_xlabel("Window length (s)")
        ax.set_ylabel("Detected boundaries (%)")
        ax.set_ylim(0, 100)
        ax.legend(frameon=False, ncol=3)
        ax.set_title("Direction-calibrated real and pseudo-boundary detections")
        _save_figure(fig, output_dir / "Figure_1_real_vs_pseudo", dpi)

    methods = detection_summary[
        detection_summary["WindowLength_sec"].eq(30)
        & detection_summary["RRThreshold"].eq(20.0)
        & detection_summary["Representation"].eq("reduced")
        & detection_summary["SearchWindow"].eq("post_only")
    ].copy()
    if not methods.empty:
        method_order = [
            "LegacyBIC_fixed6", "LegacyBIC_crossfit", "CovIC_crossfit",
            "BinSeg_L2_crossfit", "SegmentedTrend_crossfit", "CUSUM_crossfit",
            "MOSUM_crossfit", "PELT_L2_crossfit", "PELT_L1_crossfit",
            "PELT_RBF_crossfit",
        ]
        method_order = [m for m in method_order if m in set(methods["Method"])]
        fig, ax = plt.subplots(figsize=(8.5, 4.2))
        real = methods[methods["BoundaryKind"].eq("Real")].groupby("Method")["DetectionRate"].mean()
        pseudo = methods[methods["BoundaryKind"].eq("Pseudo")].groupby("Method")["DetectionRate"].mean()
        x = np.arange(len(method_order))
        ax.bar(x - 0.18, 100 * real.reindex(method_order), width=0.36, label="Real")
        ax.bar(x + 0.18, 100 * pseudo.reindex(method_order), width=0.36, label="Pseudo")
        ax.set_xticks(x, [value.replace("_crossfit", "").replace("_", "\n") for value in method_order])
        ax.set_ylabel("Detection rate (%)")
        ax.set_ylim(0, 100)
        ax.legend(frameon=False)
        ax.set_title("Method comparison, 30-s reduced representation")
        _save_figure(fig, output_dir / "Figure_2_method_comparison", dpi)

    if not comparisons.empty:
        subset = comparisons[
            comparisons["Method"].eq("CovIC_crossfit")
            & comparisons["RRThreshold"].eq(20.0)
            & comparisons["Representation"].eq("reduced")
            & comparisons["SearchWindow"].eq("post_only")
        ].copy()
        if not subset.empty:
            subset["Label"] = (
                subset["WindowLength_sec"].astype(str) + " s, "
                + subset["TransitionType"].astype(str) + ", pseudo "
                + subset["PseudoComparison"].astype(str)
            )
            subset = subset.sort_values(["WindowLength_sec", "TransitionType", "PseudoComparison"])
            fig, ax = plt.subplots(figsize=(7.0, 5.0))
            y = np.arange(len(subset))
            estimate = subset["RiskDifference"].to_numpy(dtype=float)
            low = subset["RiskDifferenceCI_Lower"].to_numpy(dtype=float)
            high = subset["RiskDifferenceCI_Upper"].to_numpy(dtype=float)
            ax.errorbar(estimate, y, xerr=np.vstack([estimate - low, high - estimate]), fmt="o", capsize=2)
            ax.axvline(0, linestyle="--", linewidth=0.8)
            ax.set_yticks(y, subset["Label"])
            ax.set_xlabel("Real minus pseudo detection probability")
            ax.set_title("Matched real-pseudo contrasts")
            _save_figure(fig, output_dir / "Figure_3_real_pseudo_risk_differences", dpi)

    if not simulation_power.empty:
        subset = simulation_power[
            simulation_power["WindowLength_sec"].eq(30)
            & simulation_power["Representation"].eq("reduced")
            & simulation_power["SearchWindow"].eq("post_only")
            & simulation_power["AffectedFraction"].eq(0.5)
            & simulation_power["ChangePosition"].eq("middle")
        ].copy()
        if not subset.empty:
            for shape in sorted(subset["ChangeShape"].unique()):
                fig, ax = plt.subplots(figsize=(6.5, 4.0))
                source = subset[subset["ChangeShape"].eq(shape)]
                for method, group in source.groupby("Method"):
                    group = group.sort_values("EffectSizeSD")
                    ax.plot(group["EffectSizeSD"], 100 * group["DetectionPower"], marker="o", label=method)
                ax.set_xlabel("Injected effect per affected dimension (SD)")
                ax.set_ylabel("Detection power (%)")
                ax.set_ylim(0, 105)
                ax.legend(frameon=False, fontsize=6, ncol=2)
                ax.set_title(f"Simulation power for {shape} changes")
                _save_figure(fig, output_dir / f"Figure_4_simulation_power_{shape}", dpi)

    if not boundary_order.empty:
        subset = boundary_order[
            boundary_order["WindowLength_sec"].eq(30)
            & boundary_order["RRThreshold"].eq(20.0)
        ]
        if not subset.empty:
            fig, ax = plt.subplots(figsize=(6.4, 3.8))
            for direction, group in subset.groupby("TransitionType"):
                group = group.sort_values("BoundaryOrder")
                ax.plot(group["BoundaryOrder"], 100 * group["DetectionRate"], marker="o", label=direction)
            ax.set_xlabel("Boundary sequence position")
            ax.set_ylabel("Detection rate (%)")
            ax.set_ylim(0, 105)
            ax.legend(frameon=False)
            ax.set_title("Primary detection by boundary position")
            _save_figure(fig, output_dir / "Figure_5_boundary_order", dpi)

    if not quality_eligibility.empty:
        summary = quality_eligibility.groupby(["WindowLength_sec", "RRThreshold"], as_index=False).agg(
            EligibleBoundaries=("Eligible", "sum")
        )
        fig, ax = plt.subplots(figsize=(6.0, 3.7))
        for window, group in summary.groupby("WindowLength_sec"):
            group = group.sort_values("RRThreshold")
            ax.plot(group["RRThreshold"], group["EligibleBoundaries"], marker="o", label=f"{window} s")
        ax.set_xlabel("Maximum RR-correction rate (%)")
        ax.set_ylabel("Eligible real boundaries")
        ax.legend(frameon=False)
        ax.set_title("ECG-quality sensitivity")
        _save_figure(fig, output_dir / "Figure_6_quality_sensitivity", dpi)

    if not feature_summary.empty:
        subset = feature_summary[
            feature_summary["WindowLength_sec"].eq(30)
            & feature_summary["RRThreshold"].eq(20.0)
            & feature_summary["BoundaryKind"].eq("Real")
        ].copy()
        if not subset.empty:
            top = subset.groupby("Feature")["MedianAbsoluteZChange"].mean().nlargest(10).index
            pivot = subset[subset["Feature"].isin(top)].pivot_table(
                index="Feature", columns="TransitionType", values="MeanZChange", aggfunc="mean"
            ).reindex(top)
            fig, ax = plt.subplots(figsize=(6.5, 4.5))
            y = np.arange(len(pivot))
            width = 0.34
            for offset, direction in zip([-width / 2, width / 2], ["NA_to_A", "A_to_NA"]):
                values = pivot.get(direction, pd.Series(index=pivot.index, dtype=float)).to_numpy(dtype=float)
                ax.barh(y + offset, values, height=width, label=direction)
            ax.axvline(0, linewidth=0.8)
            ax.set_yticks(y, pivot.index)
            ax.set_xlabel("Mean post-minus-pre change (normalized units)")
            ax.legend(frameon=False)
            ax.set_title("Magnitude and direction of feature changes")
            _save_figure(fig, output_dir / "Figure_7_feature_changes", dpi)

    if not empirical_pvalues.empty:
        subset = empirical_pvalues[
            empirical_pvalues["RRThreshold"].eq(20.0)
            & empirical_pvalues["Representation"].eq("reduced")
            & empirical_pvalues["SearchWindow"].eq("post_only")
            & empirical_pvalues["Method"].eq("CovIC")
        ]
        if not subset.empty:
            fig, ax = plt.subplots(figsize=(6.0, 3.6))
            ax.hist(subset["EmpiricalPValue"].dropna(), bins=np.linspace(0, 1, 11), edgecolor="white")
            ax.axvline(0.05, linestyle="--", linewidth=0.8)
            ax.set_xlabel("Matched empirical P value")
            ax.set_ylabel("Real boundaries")
            ax.set_title("Boundary-level evidence relative to matched pseudo controls")
            _save_figure(fig, output_dir / "Figure_8_matched_empirical_pvalues", dpi)

    timing_summary = (
        population_timing_summary.copy()
        if population_timing_summary is not None
        else pd.DataFrame()
    )
    timing_profiles = (
        population_timing_profiles.copy()
        if population_timing_profiles is not None
        else pd.DataFrame()
    )
    if not timing_profiles.empty:
        subset = timing_profiles[
            timing_profiles["WindowLength_sec"].eq(30)
            & timing_profiles["RRThreshold"].eq(20.0)
            & timing_profiles["Representation"].eq("reduced")
            & timing_profiles["Endpoint"].eq("departure_magnitude")
            & timing_profiles["SearchWindow"].eq("post_only")
        ].copy()
        if not subset.empty:
            for direction, group in subset.groupby("TransitionType", sort=False):
                group = group.sort_values("CandidateTime_sec")
                fig, ax = plt.subplots(figsize=(6.2, 3.8))
                ax.plot(
                    group["CandidateTime_sec"],
                    group["Score"],
                    marker="o",
                    label="Real participant-pooled score",
                )
                if "PseudoA_PointwiseQ95" in group:
                    ax.plot(
                        group["CandidateTime_sec"],
                        group["PseudoA_PointwiseQ95"],
                        linestyle="--",
                        label="Pseudo A pointwise 95th percentile",
                    )
                if "PseudoNA_PointwiseQ95" in group:
                    ax.plot(
                        group["CandidateTime_sec"],
                        group["PseudoNA_PointwiseQ95"],
                        linestyle=":",
                        label="Pseudo NA pointwise 95th percentile",
                    )
                ax.set_xlabel("Candidate first-new-regime window centre (s)")
                ax.set_ylabel("Participant-normalized segmented gain")
                ax.set_title(f"Participant-pooled timing profile: {direction}")
                ax.legend(frameon=False, fontsize=7)
                _save_figure(
                    fig,
                    output_dir / f"Figure_9_population_timing_profile_{direction}",
                    dpi,
                )

    if not timing_summary.empty:
        subset = timing_summary[
            timing_summary["RRThreshold"].eq(20.0)
            & timing_summary["Representation"].eq("reduced")
            & timing_summary["Endpoint"].eq("departure_magnitude")
            & timing_summary["SearchWindow"].eq("post_only")
        ].copy()
        if not subset.empty:
            fig, ax = plt.subplots(figsize=(6.4, 3.9))
            offsets = {"NA_to_A": -0.12, "A_to_NA": 0.12}
            for direction, group in subset.groupby("TransitionType", sort=False):
                group = group.sort_values("WindowLength_sec")
                x = group["WindowLength_sec"].to_numpy(dtype=float) + offsets.get(direction, 0.0)
                y = group["CandidateTime_sec"].to_numpy(dtype=float)
                low = group["CandidateTimeCI95_Lower_sec"].to_numpy(dtype=float)
                high = group["CandidateTimeCI95_Upper_sec"].to_numpy(dtype=float)
                ax.errorbar(
                    x,
                    y,
                    yerr=np.vstack([y - low, high - y]),
                    marker="o",
                    capsize=3,
                    label=direction,
                )
                validated = group["ValidatedAgainstBothPseudoControls"].fillna(False).astype(bool)
                for x_value, y_value, is_validated in zip(x, y, validated):
                    ax.annotate(
                        "validated" if is_validated else "candidate",
                        (x_value, y_value),
                        xytext=(0, 7),
                        textcoords="offset points",
                        ha="center",
                        fontsize=6,
                    )
            ax.set_xticks([30, 45, 60])
            ax.set_xlabel("Window length (s)")
            ax.set_ylabel("Candidate timing centre (s)")
            ax.set_title("Participant-pooled candidate timing by direction")
            ax.legend(frameon=False)
            _save_figure(fig, output_dir / "Figure_10_population_timing_across_windows", dpi)

    timing_simulation = (
        population_timing_simulation.copy()
        if population_timing_simulation is not None
        else pd.DataFrame()
    )
    if not timing_simulation.empty:
        subset = timing_simulation[
            timing_simulation["WindowLength_sec"].eq(30)
            & timing_simulation["RRThreshold"].eq(20.0)
            & timing_simulation["Representation"].eq("reduced")
            & timing_simulation["Endpoint"].eq("departure_magnitude")
            & timing_simulation["SearchWindow"].eq("post_only")
            & timing_simulation["ChangePosition"].eq("middle")
        ].copy()
        if not subset.empty:
            for direction, group in subset.groupby("TransitionType", sort=False):
                fig, ax = plt.subplots(figsize=(6.4, 3.9))
                grouped = (
                    group.groupby(
                        ["EffectSizeSD", "ChangeShape"],
                        as_index=False,
                    )["DetectionPower"]
                    .mean()
                )
                for shape, shape_group in grouped.groupby("ChangeShape"):
                    shape_group = shape_group.sort_values("EffectSizeSD")
                    ax.plot(
                        shape_group["EffectSizeSD"],
                        100.0 * shape_group["DetectionPower"],
                        marker="o",
                        label=shape,
                    )
                ax.set_xlabel("Injected shared change (pre-boundary SD units)")
                ax.set_ylabel("Dual-pseudo-calibrated detection power (%)")
                ax.set_ylim(0, 105)
                ax.set_title(f"Population timing simulation power: {direction}")
                ax.legend(frameon=False)
                _save_figure(
                    fig,
                    output_dir / f"Figure_11_population_timing_power_{direction}",
                    dpi,
                )


def write_readme(
    output_root: Path,
    script_name: str,
    config: object,
    workbook: Path,
) -> None:
    payload = getattr(config, "__dict__", {})
    text = f"""# Nature-oriented ECG transition and timing validation V4

Generated by `{script_name}`. The original analysis is not overwritten.

## Major methodological changes

- A participant-pooled, pseudo-calibrated shared-breakpoint model reports candidate timing for both directions and separately labels whether each timing is validated against stable A and stable NA pseudo controls.

- Only 30-, 45-, and 60-second windows are analysed.
- Symmetric A/NA participant normalization excludes the session(s) used by each boundary.
- Pseudo controls are divided into disjoint temporal blocks and cross-fitted folds.
- Thresholds and PELT penalties are calibrated separately by transition direction.
- Boundary-level empirical P values use multiple participant-matched pseudo blocks.
- Abrupt, gradual-trend, CUSUM, MOSUM, binary-segmentation, and L1/L2/RBF PELT-objective methods are compared.
- A post-only analysis is accompanied by a prespecified -30 to +60 s anticipatory sensitivity analysis.
- End-to-end leave-one-participant-out validation refits feature selection, PCA/covariance, and thresholds using the other 18 participants.

## Interpretation

All detections must be described as video-session-boundary-associated cardiovascular changes. The dataset cannot isolate anxiety from video change, attention, valence, respiration, movement, expectancy, habituation, fatigue, or orienting responses.

## Main workbook

`{workbook.name}`

## Configuration

```json
{json.dumps(payload, indent=2, default=str)}
```
"""
    (output_root / "README_RESULTS_V4.md").write_text(text, encoding="utf-8")
